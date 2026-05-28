#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.14"
# dependencies = [
#   "mammoth>=1.8",
#   "openpyxl>=3.0",
#   "pymupdf>=1.24",
#   "anthropic>=0.39",
# ]
# ///
"""Document conversion pipeline (.docx / .pdf -> .md).

Pure-by-default: every parsing/cleanup function takes inputs and returns outputs.
Filesystem and network side effects are isolated in the `run_conversion` entry
point at the bottom of this module so they're easy to spot and test.

Streams structured progress events on stdout so the UI can subscribe:

    :::progress {"phase": "parse|ocr|cleanup|cache", "current": N, "total": M}
    :::diagnostic {"code": "...", "message": "..."}

Outputs (under DOMAINS_DIR/<domain>/input/):
    policy_docs/<basename>.md
    _originals/<basename>.<ext>
    _originals/<basename>.diagnostics.json
"""

from __future__ import annotations

import base64
import dataclasses
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

CLEANUP_MODEL = "claude-sonnet-4-6"
VISION_MODEL = "claude-sonnet-4-6"

# Auto-cleanup gate thresholds (per plan).
AUTO_CLEANUP_PAGE_LIMIT = 30
AUTO_CLEANUP_TOKEN_LIMIT = 50_000

# Heuristic: average extracted chars per page below this -> treat as scanned.
SCANNED_PDF_CHARS_PER_PAGE_THRESHOLD = 50

# Rough heuristic: characters / 4 ~ token count.
CHARS_PER_TOKEN_ESTIMATE = 4


FormatLabel = Literal["docx", "pdf-text", "pdf-scanned", "csv", "xlsx", "json"]

# Structured data formats: cleanup is a document-artifact pass and doesn't
# apply to tabular or serialised data.
NO_CLEANUP_FORMATS: frozenset[str] = frozenset({"csv", "xlsx", "json"})


@dataclass
class Warning_:
    """A non-fatal issue encountered during conversion.

    Named with a trailing underscore to avoid the stdlib builtin `Warning` clash.
    """

    code: str
    detail: str
    count: int | None = None
    pages: list[int] | None = None


@dataclass
class ConversionDiagnostics:
    """Structured diagnostics written next to the original source.

    Mirrors the schema in the plan; serialized via `dataclasses.asdict`.
    """

    source: str
    source_sha256: str
    format_detected: FormatLabel
    page_count: int
    raw_markdown_bytes: int
    estimated_input_tokens: int
    cleanup_applied: bool
    cleanup_model: str | None
    warnings: list[Warning_]
    duration_ms: int


@dataclass
class ParseResult:
    """Output of the format-specific parser stage."""

    markdown: str
    page_count: int
    warnings: list[Warning_] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------


def compute_sha256(content: bytes) -> str:
    """Return hex-encoded SHA-256 of `content`."""
    hasher = hashlib.sha256()
    hasher.update(content)
    return hasher.hexdigest()


def estimate_tokens(text: str) -> int:
    """Rough token count: chars / 4 (heuristic from the plan)."""
    return max(1, len(text) // CHARS_PER_TOKEN_ESTIMATE)


def detect_format(source: Path, content: bytes) -> FormatLabel:
    """Classify the source file by extension; PDFs are further probed for text density."""
    suffix = source.suffix.lower()
    if suffix == ".docx":
        return "docx"
    if suffix == ".csv":
        return "csv"
    if suffix == ".xlsx":
        return "xlsx"
    if suffix == ".json":
        return "json"
    if suffix != ".pdf":
        raise ValueError(f"Unsupported extension: {suffix!r}")

    # PDF: probe text density.
    import fitz  # noqa: PLC0415 — heavy import only when needed

    doc = fitz.open(stream=content, filetype="pdf")
    try:
        if doc.page_count == 0:
            return "pdf-scanned"
        total_chars = 0
        for page in doc:
            total_chars += len(page.get_text("text") or "")
        average_chars = total_chars / max(1, doc.page_count)
    finally:
        doc.close()

    if average_chars < SCANNED_PDF_CHARS_PER_PAGE_THRESHOLD:
        return "pdf-scanned"
    return "pdf-text"


def parse_docx(content: bytes) -> ParseResult:
    """Convert .docx bytes to raw markdown via mammoth."""
    import mammoth  # noqa: PLC0415

    import io  # noqa: PLC0415

    result = mammoth.convert_to_markdown(io.BytesIO(content))
    warnings: list[Warning_] = []
    image_messages = [
        m for m in (result.messages or []) if "image" in str(m).lower()
    ]
    if image_messages:
        warnings.append(
            Warning_(
                code="dropped_images",
                detail="image extraction not implemented",
                count=len(image_messages),
            )
        )
    # mammoth does not surface a page count for .docx; treat the whole doc as one page.
    return ParseResult(
        markdown=result.value or "", page_count=1, warnings=warnings
    )


def _escape_table_cell(value: str) -> str:
    """Pure: make a string safe for use inside a GFM table cell."""
    return value.replace("|", "\\|").replace("\r", "").replace("\n", " ")


def rows_to_markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    """Pure: render headers + rows as a GFM pipe table."""
    escaped_headers = [_escape_table_cell(h) for h in headers]
    separator = ["-" * max(3, len(h)) for h in escaped_headers]
    lines: list[str] = [
        "| " + " | ".join(escaped_headers) + " |",
        "| " + " | ".join(separator) + " |",
    ]
    for row in rows:
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        cells = [_escape_table_cell(str(cell)) for cell in padded[: len(headers)]]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def parse_csv(content: bytes) -> ParseResult:
    """Convert CSV bytes to a single markdown table."""
    import csv  # noqa: PLC0415
    import io  # noqa: PLC0415

    text = content.decode("utf-8-sig")  # strip BOM if present
    reader = csv.reader(io.StringIO(text))
    all_rows = [row for row in reader if any(cell.strip() for cell in row)]

    warnings: list[Warning_] = []
    if not all_rows:
        return ParseResult(markdown="*(empty CSV)*\n", page_count=1, warnings=warnings)

    headers = all_rows[0]
    data_rows = all_rows[1:]

    col_counts = {len(row) for row in data_rows}
    if len(col_counts) > 1:
        warnings.append(
            Warning_(
                code="inconsistent_column_count",
                detail=f"Row widths vary: {sorted(col_counts)}",
            )
        )

    return ParseResult(
        markdown=rows_to_markdown_table(headers, data_rows) + "\n",
        page_count=1,
        warnings=warnings,
    )


def parse_xlsx(content: bytes) -> ParseResult:
    """Convert XLSX bytes to markdown — one `##`-headed section per sheet."""
    import io  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sections: list[str] = []
    warnings: list[Warning_] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        all_rows = [
            [str(cell.value) if cell.value is not None else "" for cell in row]
            for row in sheet.iter_rows()
            if any(cell.value is not None for cell in row)
        ]
        if not all_rows:
            warnings.append(
                Warning_(code="empty_sheet", detail=f"Sheet '{sheet_name}' has no data")
            )
            continue
        headers = all_rows[0]
        data_rows = all_rows[1:]
        sections.append(f"## {sheet_name}\n\n{rows_to_markdown_table(headers, data_rows)}")

    sheet_count = len(workbook.sheetnames)
    workbook.close()
    markdown = "\n\n".join(sections) + "\n" if sections else "*(no sheet data)*\n"
    return ParseResult(markdown=markdown, page_count=sheet_count, warnings=warnings)


def _json_data_to_markdown(data: Any) -> str:
    """Pure: render JSON data as markdown — table for uniform arrays, code block otherwise."""
    if (
        isinstance(data, list)
        and data
        and all(isinstance(item, dict) for item in data)
    ):
        first_keys = list(data[0].keys())
        uniform = all(list(item.keys()) == first_keys for item in data)
        all_scalar = all(
            isinstance(value, (str, int, float, bool, type(None)))
            for item in data
            for value in item.values()
        )
        if uniform and all_scalar:
            headers = [str(key) for key in first_keys]
            rows = [
                [str(item[key]) if item[key] is not None else "" for key in first_keys]
                for item in data
            ]
            return rows_to_markdown_table(headers, rows) + "\n"
    return "```json\n" + json.dumps(data, indent=2, ensure_ascii=False) + "\n```\n"


def parse_json(content: bytes) -> ParseResult:
    """Convert JSON bytes to markdown — table for uniform arrays, code block otherwise."""
    warnings: list[Warning_] = []
    try:
        data = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        warnings.append(Warning_(code="json_parse_error", detail=str(exc)))
        return ParseResult(markdown="*(unparseable JSON)*\n", page_count=1, warnings=warnings)
    return ParseResult(
        markdown=_json_data_to_markdown(data),
        page_count=1,
        warnings=warnings,
    )


def parse_pdf_text(content: bytes) -> ParseResult:
    """Convert a text-based PDF to markdown, preserving heading hierarchy."""
    import fitz  # noqa: PLC0415

    doc = fitz.open(stream=content, filetype="pdf")
    try:
        font_sizes = collect_font_sizes(doc)
        threshold_h1, threshold_h2, threshold_h3 = derive_heading_thresholds(
            font_sizes
        )
        lines: list[str] = []
        for page in doc:
            page_dict = page.get_text("dict")
            for block in page_dict.get("blocks", []):
                if block.get("type") != 0:
                    continue
                for block_line in block.get("lines", []):
                    rendered = render_pdf_line(
                        block_line,
                        threshold_h1=threshold_h1,
                        threshold_h2=threshold_h2,
                        threshold_h3=threshold_h3,
                    )
                    if rendered is not None:
                        lines.append(rendered)
            lines.append("")  # paragraph break between pages
        return ParseResult(
            markdown="\n".join(lines).strip() + "\n",
            page_count=doc.page_count,
        )
    finally:
        doc.close()


def collect_font_sizes(doc: Any) -> list[float]:
    """Collect the font size of every span in the document. Pure-ish (reads doc)."""
    sizes: list[float] = []
    for page in doc:
        page_dict = page.get_text("dict")
        for block in page_dict.get("blocks", []):
            if block.get("type") != 0:
                continue
            for block_line in block.get("lines", []):
                for span in block_line.get("spans", []):
                    size = span.get("size")
                    if isinstance(size, (int, float)):
                        sizes.append(float(size))
    return sizes


def derive_heading_thresholds(
    font_sizes: list[float],
) -> tuple[float, float, float]:
    """Return (h1, h2, h3) font-size thresholds from observed sizes.

    Strategy: take the top three distinct sizes (descending). When fewer than
    three distinct sizes exist, fall back to a single large threshold so we
    don't generate headings for body text.
    """
    if not font_sizes:
        return (float("inf"), float("inf"), float("inf"))
    distinct = sorted({round(size, 1) for size in font_sizes}, reverse=True)
    body_size = max(distinct[-1], 1.0)
    candidates = [size for size in distinct if size > body_size * 1.1]
    if len(candidates) < 3:
        # Pad with infinity so missing tiers never match.
        padded = candidates + [float("inf")] * (3 - len(candidates))
        return (padded[0], padded[1], padded[2])
    return (candidates[0], candidates[1], candidates[2])


def render_pdf_line(
    block_line: dict[str, Any],
    *,
    threshold_h1: float,
    threshold_h2: float,
    threshold_h3: float,
) -> str | None:
    """Render one block-line as markdown (heading + body). None if empty."""
    spans = block_line.get("spans", [])
    if not spans:
        return None
    text = "".join(span.get("text", "") for span in spans).strip()
    if not text:
        return None
    max_size = max(
        (float(span.get("size", 0.0)) for span in spans), default=0.0
    )
    if max_size >= threshold_h1:
        return f"# {text}"
    if max_size >= threshold_h2:
        return f"## {text}"
    if max_size >= threshold_h3:
        return f"### {text}"
    return text


def emit_progress(
    phase: str, current: int, total: int, **extra: Any
) -> None:
    """Side-effecting: write a `:::progress` line to stdout."""
    payload: dict[str, Any] = {
        "phase": phase,
        "current": current,
        "total": total,
    }
    payload.update(extra)
    sys.stdout.write(f":::progress {json.dumps(payload)}\n")
    sys.stdout.flush()


def emit_diagnostic(code: str, **extra: Any) -> None:
    """Side-effecting: write a `:::diagnostic` line to stdout."""
    payload: dict[str, Any] = {"code": code}
    payload.update(extra)
    sys.stdout.write(f":::diagnostic {json.dumps(payload)}\n")
    sys.stdout.flush()


def parse_pdf_scanned(
    content: bytes, *, anthropic_client: Any
) -> ParseResult:
    """OCR a scanned PDF one page at a time via Claude vision."""
    import fitz  # noqa: PLC0415

    doc = fitz.open(stream=content, filetype="pdf")
    try:
        total_pages = doc.page_count
        page_markdowns: list[str] = []
        warnings: list[Warning_] = []
        low_confidence_pages: list[int] = []
        for page_index in range(total_pages):
            emit_progress("ocr", page_index + 1, total_pages)
            page = doc[page_index]
            pixmap = page.get_pixmap(dpi=200)
            png_bytes = pixmap.tobytes("png")
            page_markdown, confident = ocr_page_with_claude(
                png_bytes, anthropic_client=anthropic_client
            )
            page_markdowns.append(page_markdown)
            if not confident:
                low_confidence_pages.append(page_index + 1)
        if low_confidence_pages:
            warnings.append(
                Warning_(
                    code="low_ocr_confidence",
                    detail="Claude flagged page as partially unreadable",
                    pages=low_confidence_pages,
                )
            )
        return ParseResult(
            markdown="\n\n".join(page_markdowns).strip() + "\n",
            page_count=total_pages,
            warnings=warnings,
        )
    finally:
        doc.close()


_VISION_PROMPT = (
    "Transcribe this scanned policy document page to markdown. "
    "Preserve numbered sections, headings, lists, and tables exactly as they appear. "
    "Drop page numbers, running headers/footers, and decorative graphics. "
    "If any portion of the page is unreadable, prepend the response with the literal "
    "token UNREADABLE_PORTION on its own line; otherwise output only the markdown."
)


def ocr_page_with_claude(
    png_bytes: bytes, *, anthropic_client: Any
) -> tuple[str, bool]:
    """Send one page image to Claude vision and return (markdown, confident)."""
    encoded = base64.standard_b64encode(png_bytes).decode("ascii")
    response = anthropic_client.messages.create(
        model=VISION_MODEL,
        max_tokens=4096,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": encoded,
                        },
                    },
                    {"type": "text", "text": _VISION_PROMPT},
                ],
            }
        ],
    )
    text_blocks = [
        block.text for block in response.content if hasattr(block, "text")
    ]
    full_text = "\n".join(text_blocks).strip()
    confident = True
    if full_text.startswith("UNREADABLE_PORTION"):
        confident = False
        full_text = full_text.removeprefix("UNREADABLE_PORTION").lstrip("\n")
    return full_text, confident


_CLEANUP_PROMPT = (
    "You are cleaning up the output of an automated document-to-markdown converter "
    "for a policy document. Your job: preserve all policy content (numbered sections, "
    "definitions, citations, formulas, tables, lists) while removing artifacts:\n"
    "- Page numbers (e.g. lines that are just '12' or 'Page 12 of 50')\n"
    "- Running headers/footers that repeat across pages\n"
    "- Tables of contents (the actual TOC, not section headings)\n"
    "- Binary/encoded noise (long base64-looking strings, control characters)\n"
    "- Stray whitespace and broken line wrapping inside paragraphs\n\n"
    "Repair heading levels where the parser misclassified them. Normalize list and "
    "table markdown. NEVER paraphrase or summarize content. Output ONLY the cleaned "
    "markdown, no preamble.\n\n"
    "Input:\n"
)


def _cleanup_via_sdk(raw_markdown: str, *, anthropic_client: Any) -> str:
    """Run cleanup via the Anthropic Python SDK (requires ANTHROPIC_API_KEY)."""
    response = anthropic_client.messages.create(
        model=CLEANUP_MODEL,
        max_tokens=8192,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": _CLEANUP_PROMPT + raw_markdown}
                ],
            }
        ],
    )
    text_blocks = [
        block.text for block in response.content if hasattr(block, "text")
    ]
    return "\n".join(text_blocks).strip() + "\n"


def _cleanup_via_cli(raw_markdown: str) -> str:
    """Run cleanup via the `claude` CLI using the authenticated session."""
    result = subprocess.run(
        ["claude", "-p", "--output-format", "text", _CLEANUP_PROMPT + raw_markdown],
        capture_output=True,
        text=True,
        check=True,
    )
    return result.stdout.strip() + "\n"


def cleanup_markdown(raw_markdown: str, *, anthropic_client: Any | None) -> str:
    """Run a single Claude cleanup pass, using the SDK client or CLI fallback."""
    if anthropic_client is not None:
        return _cleanup_via_sdk(raw_markdown, anthropic_client=anthropic_client)
    return _cleanup_via_cli(raw_markdown)


def should_auto_run_cleanup(
    *, page_count: int, raw_markdown: str
) -> bool:
    """Pure: return True when the doc is under the auto-cleanup thresholds."""
    if page_count > AUTO_CLEANUP_PAGE_LIMIT:
        return False
    estimated_tokens = estimate_tokens(raw_markdown)
    return estimated_tokens <= AUTO_CLEANUP_TOKEN_LIMIT


def diagnostics_dict(diagnostics: ConversionDiagnostics) -> dict[str, Any]:
    """Pure: convert dataclass to JSON-serializable dict, dropping None fields."""
    raw = dataclasses.asdict(diagnostics)
    raw["warnings"] = [
        {key: value for key, value in warning.items() if value is not None}
        for warning in raw["warnings"]
    ]
    return raw


def basename_without_extension(source: Path) -> str:
    """Return `source.stem` after sanitizing characters that break shell scripts."""
    sanitized = re.sub(r"[^A-Za-z0-9._-]+", "-", source.stem).strip("-.")
    return sanitized or "document"


# ---------------------------------------------------------------------------
# Side-effecting orchestration
# ---------------------------------------------------------------------------


def find_anthropic_client() -> Any | None:
    """Build an Anthropic SDK client if the SDK is importable and key set."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return None
    try:
        import anthropic  # noqa: PLC0415
    except ImportError:
        return None
    return anthropic.Anthropic(api_key=api_key)


def claude_cli_available() -> bool:
    """Return True when the `claude` CLI is on PATH (authenticated session usable)."""
    return shutil.which("claude") is not None


def project_paths(domain: str, basename: str, extension: str) -> dict[str, Path]:
    """Pure-ish: derive output paths under DOMAINS_FULLPATH/<domain>/input/."""
    domains_dir = Path(os.environ["DOMAINS_FULLPATH"])
    base = domains_dir / domain / "input"
    return {
        "policy_dir": base / "policy_docs",
        "originals_dir": base / "_originals",
        "md_target": base / "policy_docs" / f"{basename}.md",
        "original_target": base / "_originals" / f"{basename}{extension}",
        "diagnostics_target": base
        / "_originals"
        / f"{basename}.diagnostics.json",
    }


def hash_cache_hit(
    *, source_sha: str, original_target: Path
) -> bool:
    """Return True when the cached original has the same SHA as the source."""
    if not original_target.exists():
        return False
    cached_sha = compute_sha256(original_target.read_bytes())
    return cached_sha == source_sha


def write_outputs(
    *,
    paths: dict[str, Path],
    markdown: str,
    source_content: bytes,
    diagnostics: ConversionDiagnostics,
) -> None:
    """Side-effecting: write md + original + diagnostics to disk."""
    paths["policy_dir"].mkdir(parents=True, exist_ok=True)
    paths["originals_dir"].mkdir(parents=True, exist_ok=True)
    paths["md_target"].write_text(markdown, encoding="utf-8")
    paths["original_target"].write_bytes(source_content)
    paths["diagnostics_target"].write_text(
        json.dumps(diagnostics_dict(diagnostics), indent=2) + "\n",
        encoding="utf-8",
    )


def run_conversion(
    *,
    domain: str,
    source_path: Path,
    force_cleanup: bool,
    no_cleanup: bool,
) -> int:
    """Top-level orchestration. Returns the process exit code."""
    if not source_path.is_file():
        sys.stderr.write(f"Source file not found: {source_path}\n")
        return 2

    source_content = source_path.read_bytes()
    source_sha = compute_sha256(source_content)
    basename = basename_without_extension(source_path)
    extension = source_path.suffix.lower()
    paths = project_paths(domain, basename, extension)

    if hash_cache_hit(source_sha=source_sha, original_target=paths["original_target"]):
        if paths["md_target"].exists():
            emit_progress("cache", 1, 1, basename=basename)
            sys.stdout.write(
                f"Cache hit: reusing existing {paths['md_target'].name}\n"
            )
            return 0

    started_at = time.monotonic()
    format_label = detect_format(source_path, source_content)
    emit_progress("parse", 0, 1, format=format_label)

    anthropic_client = find_anthropic_client()

    if format_label == "docx":
        parse_result = parse_docx(source_content)
    elif format_label == "pdf-text":
        parse_result = parse_pdf_text(source_content)
    elif format_label == "csv":
        parse_result = parse_csv(source_content)
    elif format_label == "xlsx":
        parse_result = parse_xlsx(source_content)
    elif format_label == "json":
        parse_result = parse_json(source_content)
    else:
        # Scanned PDFs require the Anthropic vision API — image bytes are sent
        # page-by-page and there is no equivalent claude CLI path.
        # Local OCR alternatives (marker-pdf, pymupdf4llm + rapidocr-onnxruntime,
        # docling) could serve as fallbacks but involve 1–3 GB model downloads and
        # lower quality on dense policy documents; not implemented for now.
        if anthropic_client is None:
            sys.stderr.write(
                "Scanned PDF requires ANTHROPIC_API_KEY and the `anthropic` SDK.\n"
            )
            return 3
        parse_result = parse_pdf_scanned(
            source_content, anthropic_client=anthropic_client
        )

    emit_progress("parse", 1, 1, format=format_label)

    raw_markdown = parse_result.markdown
    estimated_tokens = estimate_tokens(raw_markdown)
    cleanup_decision_auto = should_auto_run_cleanup(
        page_count=parse_result.page_count, raw_markdown=raw_markdown
    )

    cleanup_applied = False
    final_markdown = raw_markdown

    cleanup_via_cli = anthropic_client is None and claude_cli_available()
    if format_label in NO_CLEANUP_FORMATS:
        no_cleanup = True
    if no_cleanup:
        cleanup_applied = False
    elif force_cleanup or cleanup_decision_auto:
        if anthropic_client is None and not claude_cli_available():
            # Don't fail; just record a warning and keep raw markdown.
            parse_result.warnings.append(
                Warning_(
                    code="cleanup_skipped_no_api_key",
                    detail="ANTHROPIC_API_KEY not set or anthropic SDK missing, and `claude` CLI not found",
                )
            )
        else:
            emit_progress("cleanup", 0, 1)
            final_markdown = cleanup_markdown(
                raw_markdown, anthropic_client=anthropic_client
            )
            cleanup_applied = True
            emit_progress("cleanup", 1, 1)
    else:
        emit_diagnostic(
            "confirm_cleanup_required",
            page_count=parse_result.page_count,
            estimated_input_tokens=estimated_tokens,
            message=(
                f"Document exceeds auto-cleanup thresholds "
                f"(pages={parse_result.page_count}, tokens~={estimated_tokens}). "
                "Re-run with --force-cleanup to apply cleanup."
            ),
        )
        return 4

    duration_ms = int((time.monotonic() - started_at) * 1000)
    diagnostics = ConversionDiagnostics(
        source=source_path.name,
        source_sha256=source_sha,
        format_detected=format_label,
        page_count=parse_result.page_count,
        raw_markdown_bytes=len(raw_markdown.encode("utf-8")),
        estimated_input_tokens=estimated_tokens,
        cleanup_applied=cleanup_applied,
        cleanup_model=(None if not cleanup_applied else ("claude-cli" if cleanup_via_cli else CLEANUP_MODEL)),
        warnings=parse_result.warnings,
        duration_ms=duration_ms,
    )

    write_outputs(
        paths=paths,
        markdown=final_markdown,
        source_content=source_content,
        diagnostics=diagnostics,
    )

    sys.stdout.write(
        f"Wrote {paths['md_target'].relative_to(Path(os.environ['DOMAINS_FULLPATH']))}\n"
    )
    return 0


# ---------------------------------------------------------------------------
# CLI entry point (importable for tests)
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    """Command-line entry. Mirrors xlator.py argparse style minimally."""
    import argparse  # noqa: PLC0415

    parser = argparse.ArgumentParser(
        prog="xlator convert-doc",
        description="Convert a .docx, .pdf, .csv, .xlsx, or .json file into markdown for indexing.",
    )
    parser.add_argument("domain", help="Domain name (e.g. snap, ak_doh)")
    parser.add_argument("source", help="Path to source file (.docx, .pdf, .csv, .xlsx, .json)")
    parser.add_argument(
        "--force-cleanup",
        action="store_true",
        help="Run cleanup even when the doc exceeds auto-cleanup thresholds.",
    )
    parser.add_argument(
        "--no-cleanup",
        action="store_true",
        help="Skip cleanup entirely (used by hermetic tests).",
    )
    args = parser.parse_args(argv)

    if args.force_cleanup and args.no_cleanup:
        parser.error("--force-cleanup and --no-cleanup are mutually exclusive")

    return run_conversion(
        domain=args.domain,
        source_path=Path(args.source).expanduser().resolve(),
        force_cleanup=args.force_cleanup,
        no_cleanup=args.no_cleanup,
    )


if __name__ == "__main__":
    sys.exit(main())


# Convenience for tests: re-export the dataclasses without leaking internals.
__all__ = [
    "ConversionDiagnostics",
    "ParseResult",
    "Warning_",
    "basename_without_extension",
    "compute_sha256",
    "derive_heading_thresholds",
    "detect_format",
    "diagnostics_dict",
    "estimate_tokens",
    "hash_cache_hit",
    "main",
    "parse_csv",
    "parse_docx",
    "parse_json",
    "parse_pdf_text",
    "parse_xlsx",
    "project_paths",
    "rows_to_markdown_table",
    "run_conversion",
    "should_auto_run_cleanup",
]

